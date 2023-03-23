import openpyxl

# Открытие файла
workbook = openpyxl.load_workbook('расписание 13.03.2023.xlsx')

# Получение нужного листа
worksheet = workbook['ИСиП']

# Создание списка, который будет содержать данные из нужных строк
data = []
for row in worksheet.iter_rows(min_row=1):
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    data.append(row_data)

# Поиск индекса строки, содержащей нужный текст
start_row = None
for i, row in enumerate(data):
    if 'День' in str(row[0]):
        start_row = i + 1
        break

# Вывод данных столбцами, пропуская None
if start_row is not None:
    for row in zip(*data[start_row:]):
        for cell in row:
            if cell is not None:
                print(cell, end='\t')
        print()

# Создание нового файла Excel
workbook = openpyxl.Workbook()

# Добавление листа
worksheet = workbook.active
worksheet.title = 'Мой лист'

# Заполнение листа данными
for row in data:
    worksheet.append(row)

# Сохранение файла
workbook.save('lansha.xlsx')