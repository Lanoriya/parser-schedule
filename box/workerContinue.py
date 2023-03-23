import openpyxl

# Открытие файла
workbook = openpyxl.load_workbook('lansha.xlsx')

# Получение нужного листа
worksheet = workbook.active

# Создание нового файла и листа
new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active  # Изменено на new_workbook.active

# Считывание и запись данных в новый файл
for row in worksheet.iter_rows():
  new_row_data = []
  for i, cell in enumerate(row):
    if i in [0, 5, 6]:  # оставляем только первый, шестой и седьмой столбцы
      new_row_data.append(cell.value)
  new_worksheet.append(new_row_data)

# Сохранение нового файла
new_workbook.save('NewWorker.xlsx')
