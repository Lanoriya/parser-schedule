import requests
from bs4 import BeautifulSoup
import re
import os
import gdown
import time
import openpyxl


while True:
    # получаем содержимое страницы
    url = "https://kcpt72.ru/schedule/"
    response = requests.get(url)
    content = response.content

    # парсим страницу с помощью BeautifulSoup
    soup = BeautifulSoup(content, 'html.parser')

    for filename in os.listdir():
        if filename.endswith('.xlsx') or filename.endswith('.docx'):
            os.remove(filename)
    # находим все ссылки в блоке sheulde-content и скачиваем файлы
    for link in soup.select('.sheulde-content a'):
        href = link['href']
        if href.startswith('https://docs.google.com/spreadsheets/'):
            file_id = re.search('/d/([A-Za-z0-9_-]+)', href).group(1)
            download_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
            output_file = f"{file_id}.xlsx"
            time.sleep(3)
            gdown.download(download_url, output_file, quiet=False)

            # Переименовываем файл
            response = requests.get(href)
            soup = BeautifulSoup(response.content, 'html.parser')
            span = soup.find('span', {'class': 'docs-title-input-label-inner'})
            if span:
                new_name = span.text.strip().replace("xlsx", "")
                os.rename(output_file, f"{new_name}.xlsx")
                workbook = openpyxl.load_workbook(f'{new_name}.xlsx')
                worksheet = workbook['ИСиП']
                data = []
                for row in worksheet.iter_rows(min_row=1):
                    row_data = []
                    for cell in row:
                        row_data.append(cell.value)
                    data.append(row_data)

                # Поиск индекса строки, содержащей нужный текст
                start_row = None
                for i, row in enumerate(data):
                    if 'День' in str(row[0]):
                        start_row = i + 1
                        break

                # Вывод данных столбцами, пропуская None
                if start_row is not None:
                    for row in zip(*data[start_row:]):
                        for cell in row:
                            if cell is not None:
                                print(cell, end='\t')
                        print()

                # Создание нового файла Excel
                workbook = openpyxl.Workbook()

                # Добавление листа
                worksheet = workbook.active
                worksheet.title = 'Мой лист'

                # Заполнение листа данными
                for row in data:
                    worksheet.append(row)

                # Сохранение файла
                workbook.save(f'{new_name}-reworked.xlsx')
                workbook = openpyxl.load_workbook(f'{new_name}-reworked.xlsx')

                # Получение нужного листа
                worksheet = workbook.active

                # Создание нового файла и листа
                new_workbook = openpyxl.Workbook()
                new_worksheet = new_workbook.active  # Изменено на new_workbook.active

                # Считывание и запись данных в новый файл
                for row in worksheet.iter_rows():
                    new_row_data = []
                    for i, cell in enumerate(row):
                        if i in [0, 5, 6]:  # оставляем только первый, шестой и седьмой столбцы
                            new_row_data.append(cell.value)
                    new_worksheet.append(new_row_data)

                # Сохранение нового файла
                new_workbook.save(f'{new_name}-accept.xlsx')
        elif href.startswith('https://docs.google.com/document/d/'):
            file_id = re.search('/d/([A-Za-z0-9_-]+)/edit', href).group(1)
            download_url = f"https://docs.google.com/document/d/{file_id}/export?format=docx"
            output_file = f"{file_id}.docx"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
            time.sleep(3)
            response = requests.get(download_url, headers=headers)
            with open(output_file, 'wb') as f:
                f.write(response.content)

            # Переименовываем файл
            response = requests.get(href)
            soup = BeautifulSoup(response.content, 'html.parser')
            span = soup.find('span', {'class': 'docs-title-input-label-inner'})
            if span:
                new_name = span.text.strip()
                os.rename(output_file, f"{new_name}.docx")
            # Задержка на 2 часа перед следующим обновлением
    time.sleep(7200)
